from openpyxl import *

from config import Config
from util import check_file_exist, log, check_duration
import numpy as np


class Excel:
    DEFAULT_SAVE_PATH = "sample.xlsx"
    DEFAULT_SHEET_NAME = "sheet"

    def __init__(self):
        self.wb = None
        self.ws = None
        self.save_path = None

    def open(self, path=DEFAULT_SAVE_PATH):
        if check_file_exist(path):
            self.wb = load_workbook(path)
        else:
            self.wb = Workbook()
        self.save_path = path

    def close(self):
        self.wb.close()

    def write_data(self, **kwargs):
        device = kwargs["device"] if "data" in kwargs.keys() else None
        data = kwargs["data"] if "data" in kwargs.keys() else None
        title = kwargs["title"] if kwargs["title"] is not None else Excel.DEFAULT_SHEET_NAME
        if data is None or not isinstance(data, dict):
            return
        self.ws = self.wb.create_sheet(title, 0)  # create sheet

        if device:
            self.ws.cell(row=1, column=1, value="system version:")  # write system version
            self.ws.cell(row=1, column=2, value=device.system_version)
            self.ws.cell(row=2, column=1, value="app version:")  # write app version
            self.ws.cell(row=2, column=2, value=device.app_version)

        # write app
        app_data = data["app"]
        app_start_row = 3 if device else 1
        app_start_column = 1
        self.ws.cell(row=app_start_row, column=app_start_column, value="app")
        end_column = self.__write_item(app_data, "app_log", app_start_row + 1, app_start_column)  # write app data

        # write hal
        hal_data = data["hal"]
        hal_start_row = 3 if device else 1
        hal_start_column = app_start_column + end_column
        self.ws.cell(row=app_start_row, column=hal_start_column, value="hal")
        self.__write_item(hal_data, "hal_log", hal_start_row + 1, hal_start_column)  # write hal data

        print("output in %s" % os.path.abspath(self.save_path))
        end_column = self.wb.save(self.save_path)

    def __write_item(self, data, data_type, start_row, start_column):
        type_title_row = start_row
        type_title_column = start_column

        for i in xrange(len(data.keys())):
            key = data.keys()[i]

            type_title = Config.get_title(key, data_type)
            self.ws.cell(row=type_title_row, column=type_title_column, value=type_title)  # write header

            item = data.get(key)
            if isinstance(item, dict):
                mode_titles = item.keys()

                mode_title_row = start_row + 1
                mode_title_column = type_title_column

                for mode in mode_titles:
                    self.ws.cell(row=mode_title_row, column=mode_title_column, value=mode)  # write header
                    log("write item, mode : %s" % mode)
                    sum_value = 0
                    time_data = item.get(mode)
                    data_row = mode_title_row + 1
                    if isinstance(time_data, list):
                        for j in xrange(len(time_data)):
                            row = mode_title_row + j
                            column = mode_title_column
                            value = check_duration(time_data[j])
                            self.ws.cell(row=row, column=column, value=value)  # write data
                            sum_value += value
                    if len(time_data) > 0:
                        self.ws.cell(row=data_row + len(time_data), column=mode_title_column,
                                     value="svg:")  # write avg data
                        self.ws.cell(row=data_row + len(time_data) + 1, column=mode_title_column,
                                     value=sum_value / len(time_data))  # write avg data
                    else:
                        log("write item error, mode : %s" % mode)
                    mode_title_column = mode_title_column + 1

                type_title_column = type_title_column + len(mode_titles)
            elif isinstance(item, list):
                sum_value = 0
                data_row = start_row + 1
                for j in xrange(len(item)):
                    row = data_row + j
                    column = type_title_column
                    value = check_duration(item[j])
                    sum_value += value
                    self.ws.cell(row=row, column=column, value=value)  # write data
                if len(item) > 0:
                    self.ws.cell(row=data_row + len(item), column=type_title_column,
                                 value="svg:")  # write avg data
                    self.ws.cell(row=data_row + len(item) + 1, column=type_title_column,
                                 value=sum_value / len(item))  # write avg data
                type_title_column = type_title_column + 1
            else:
                log("write item error, key : %s" % key)
        return type_title_column
