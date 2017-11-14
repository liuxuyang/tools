from openpyxl import *

from config import Config
from util import check_file_exist, log


class Excel:
    DEFAULT_SAVE_PATH = "sample.xlsx"
    DEFAULT_SHEET_NAME = "sheet"

    def __init__(self):
        self.wb = None
        self.ws = None
        self.save_path = None

    def open(self, path=DEFAULT_SAVE_PATH):
        if path and check_file_exist(path):
            self.wb = load_workbook(path)
            self.save_path = path
        else:
            self.wb = Workbook()
            self.save_path = Excel.DEFAULT_SAVE_PATH

    def close(self):
        self.wb.close()

    def write_data(self, **kwargs):
        data = kwargs["data"] if "data" in kwargs.keys() else None
        title = kwargs["title"] if kwargs["title"] is not None else Excel.DEFAULT_SHEET_NAME
        if data is None or not isinstance(data, dict):
            return
        self.ws = self.wb.create_sheet(title, 0)  # create sheet
        for i in xrange(len(data.keys())):
            title = Config.get_title("hal_log", data.keys()[i])
            self.ws.cell(row=1, column=i + 1, value=title)  # write header
            key = data.keys()[i]
            column_data = data.get(key)
            for j in xrange(len(column_data)):
                value = column_data[j]
                self.ws.cell(row=j + 1 + 1, column=i + 1, value=value)  # write data
        print("output in %s" % os.path.abspath(self.save_path))
        self.wb.save(self.save_path)
